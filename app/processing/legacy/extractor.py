import re
from typing import Dict
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Optional PDF support
try:
    import pdfplumber
except ImportError:
    pdfplumber = None


# ======================================================
# REGEX SETUP (USER-PROVIDED – UNCHANGED LOGIC)
# ======================================================

def _setup_regex_patterns() -> Dict[str, re.Pattern]:
    patterns = {}
    patterns['single'] = re.compile(
        r'(?:\(|\b)(Figure|Fig\.?|Table|Tab\.?|Box|Image|Img\.?|Photo|Illustration)\.?\s*'
        r'([0-9]+(?:[.\-][0-9]+)*)([A-Za-z]?)(?:\)|\b)',
        re.IGNORECASE
    )
    patterns['range'] = re.compile(
        r'(?:\(|\b)(Figures?|Figs?\.?|Tables?|Tabs?\.?|Boxes?|Images?|Imgs?\.?|Photos?|Illustrations?)\.?\s+'
        r'([0-9]+(?:[\.\-][0-9]+)+)([A-Za-z]?)\s*'
        r'(?:to|through|–|—|-)\s*'
        r'([0-9]+(?:[\.\-][0-9]+)*)([A-Za-z]?)(?:\)|\b)',
        re.IGNORECASE
    )
    patterns['and'] = re.compile(
        r'(?:\(|\b)(Figures?|Figs?\.?|Tables?|Tabs?\.?|Boxes?|Images?|Imgs?\.?|Photos?|Illustrations?)\.?\s+'
        r'([0-9]+(?:[\.\-][0-9]+)+)([A-Za-z]?)\s+'
        r'(?:and|&)\s*'
        r'([0-9]+(?:[\.\-][0-9]+)*)([A-Za-z]?)(?:\)|\b)',
        re.IGNORECASE
    )
    # Catch-all for unnumbered items or those that didn't match strict numbering
    patterns['unnumbered'] = re.compile(
        r'(?:\(|\b)(Figure|Fig\.?|Table|Tab\.?|Box|Image|Img\.?|Photo|Illustration)\.?(?:\s+|$)',
        re.IGNORECASE
    )
    return patterns


CAPTION_PATTERNS = _setup_regex_patterns()


# ======================================================
# CHAPTER DETECTION
# ======================================================

CHAPTER_REGEX = re.compile(
    r"(?i)^(chapter\s+\d+|\d+\.\s+[A-Z][A-Za-z ].+)"
)


# ======================================================
# CREDIT DETECTION
# ======================================================

CREDIT_KEYWORDS = [
    r"sources?:\s*",
    r"adapted\s+(?:with\s+permission\s+)?(?:from|of)",
    r"modified\s+(?:with\s+permission\s+)?(?:from|of)",
    r"based\s+on",
    r"reprinted\s+(?:with\s+permission\s+)?(?:from|of)",
    r"reproduced\s+(?:with\s+permission\s+)?(?:from|of)",
    r"data\s+from",
    r"with\s+permission",
    r"published\s+with\s+permission",
    r"copyright",
    r"©",
    r"courtesy\s+of",
    r"images?\s+courtesy",
    r"photo\s+credit",
    r"illustration\s+by",
    r"illustrated\s+by",
    r"shutterstock",
    r"getty",
    r"retrieved\s+from",
    r"accessed\s+from",
    r"https?://",
    r"doi\.org/"
]

CREDIT_REGEX = re.compile(r"(?i)(" + "|".join(CREDIT_KEYWORDS) + ")")


EDITORIAL_EXCLUDE_REGEX = re.compile(
    r"""(?i)^(
        this\s+is|
        please\s+advise|
        we\s+could|
        it\s+would|
        part\s+of\s+this
    )""",
    re.VERBOSE
)


# ======================================================
# PERMISSION RISK
# ======================================================

PERMISSION_RISK_REGEX = re.compile(
    r"""(?i)(
        adapted\s+(?:with\s+permission\s+)?from|
        modified\s+(?:with\s+permission\s+)?from|
        based\s+on|
        reproduced\s+(?:with\s+permission\s+)?from|
        reprinted\s+(?:with\s+permission\s+)?from|
        courtesy\s+of|
        copyright|
        ©|
        from\s+another\s+book|
        journal|
        press|
        university|
        https?://|
        doi\.org/|
        sources?[\.\s:]*
    )""",
    re.VERBOSE
)

CAPTION_START_REGEX = re.compile(
    r'^\s*(Figure|Fig\.?|Table|Tab\.?|Box|Image|Img\.?|Photo|Illustration)',
    re.IGNORECASE
)
# ======================================================
# CREDIT EXTRACTION
# ======================================================

def extract_credit_sentence(text):
    """
    Extract full formal credit blocks without truncation.
    Handles initials, degrees, locations, and multiple credits.
    Enhanced to capture full source blocks including URLs and DOIs.
    """
    if not CREDIT_REGEX.search(text):
        return None

    text = re.sub(r"\s+", " ", text).strip()

    credits = []

    # Special handling for "Sources:" at paragraph start (must be at very start)
    # Special handling for "Source(s):" at paragraph start
    if re.match(r'^sources?[\s:]', text, re.IGNORECASE):
        # Capture entire paragraph for sources block
        credit_block = text
        # Stop at obvious breaks (note, please advise, etc)
        credit_block = re.split(
            r'(?i)(please advise|either redraw|if using the original)',
            credit_block
        )[0]
        credits.append(credit_block.strip())
    else:
        # Only capture other keywords if NOT in a "Note:" or similar section
        # Skip if this is a note/disclaimer paragraph
        if not re.match(r'(?i)^note:', text):
            # Original logic for other credit keywords
            for match in CREDIT_REGEX.finditer(text):
                start = match.start()
                # Capture preceding parenthesis if present
                if start > 0 and text[start-1] in '([':
                    start -= 1
                
                credit_block = text[start:]
                
                # Stop at obvious editorial instructions
                credit_block = re.split(
                    r'(?i)(please advise|either redraw|if using the original)',
                    credit_block
                )[0]
                
                credits.append(credit_block.strip())

    if not credits:
        return None

    # Post-process credits to clean up common prefixes
    cleaned_credits = []
    for c in credits:
        c = c.strip()
        # Remove leading parenthesis
        if c.startswith('('):
            c = c[1:].strip()
        
        # Remove "Used" prefix variants (Used, Used:, Used-, etc)
        c = re.sub(r'^used\b[\s:\-\.]*', '', c, flags=re.IGNORECASE)
            
        cleaned_credits.append(c)
    credits = cleaned_credits

    # De-duplicate: remove exact duplicates and overlapping URL/DOI patterns
    seen = set()
    final = []
    
    for c in credits:
        # Skip if exact match already seen
        if c in seen:
            continue
        
        # Check if this is just a URL/DOI that's already part of a previous longer credit
        is_substring = False
        for prev in final:
            if c in prev and len(c) < len(prev):
                is_substring = True
                break
        
        if not is_substring:
            final.append(c)
            seen.add(c)

    # Join and remove duplicate URLs/DOIs from the final string
    result = " ".join(final)
    
    # Remove consecutive duplicate DOIs and URLs
    # Match https://doi.org/... and https://doi.org/... patterns and keep only one
    result = re.sub(r'(https?://doi\.org/[^\s]+)(\s+\1)+', r'\1', result)
    result = re.sub(r'(https?://[^\s]+)(\s+\1)+', r'\1', result)
    # Also remove standalone doi.org/... if https version exists
    result = re.sub(r'(https://doi\.org/[^\s]+)\s+doi\.org/([^\s]+)', r'\1', result)
    
    return result.strip()



def needs_permission(caption, credit):
    # If credit exists, check if it contains external publication markers
    if credit:
        # Check for external sources (URLs, DOIs, publication references)
        if PERMISSION_RISK_REGEX.search(credit):
            return "YES"
        # Any credit line from external source (contains reference info) needs permission
        if any(marker in credit.lower() for marker in ['sources', 'retrieved', 'accessed', 'doi', 'http']):
            return "YES"
        return "NO"
    if PERMISSION_RISK_REGEX.search(caption):
        return "YES"
    return "NO"


# ======================================================
# TEXT EXTRACTION
# ======================================================

def extract_text_from_docx(path):
    doc = Document(path)
    return [p.text.strip() for p in doc.paragraphs if p.text.strip()]


def extract_text_from_pdf(path):
    if not pdfplumber:
        raise RuntimeError("pdfplumber not installed")
    paras = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            paras.extend([l.strip() for l in text.split("\n") if l.strip()])
    return paras


# ======================================================
# CAPTION MATCHING (USING YOUR REGEX)
# ======================================================

def match_caption(paragraph):
    # MUST start like a caption, not mid-sentence citation
    if not CAPTION_START_REGEX.match(paragraph):
        return None, None

    for ptype, regex in CAPTION_PATTERNS.items():
        m = regex.match(paragraph)
        if m:
            return ptype, m

    return None, None



def normalize_item_type(raw):
    raw = raw.lower()
    if "fig" in raw:
        return "Figure"
    if "tab" in raw:
        return "Table"
    if "box" in raw:
        return "Box"
    if "image" in raw or "img" in raw or "photo" in raw or "illustration" in raw:
        return "Figure"
    if "exhibit" in raw:
        return "Exhibit"
    if "appendix" in raw:
        return "Appendix"
    return "Case Study"


# ======================================================
# MAIN EXTRACTION
# ======================================================

def extract_from_file(path):
    paragraphs = (
        extract_text_from_pdf(path)
        if path.lower().endswith(".pdf")
        else extract_text_from_docx(path)
    )

    results = []
    current_chapter = ""

    for i, para in enumerate(paragraphs):

        if CHAPTER_REGEX.match(para):
            current_chapter = para
            continue

        ptype, match = match_caption(para)
        if not match:
            continue

        raw_type = match.group(1)
        item_type = normalize_item_type(raw_type)

        if ptype == "single":
            item_no = match.group(2) + (match.group(3) or "")
        elif ptype == "range":
            item_no = f"{match.group(2)}{match.group(3) or ''}–{match.group(4)}{match.group(5) or ''}"
        elif ptype == "and":
            item_no = f"{match.group(2)}{match.group(3) or ''} & {match.group(4)}{match.group(5) or ''}" 
        else:
            # Unnumbered
            item_no = ""

        # Extract chapter number from item number (e.g., "1.1" -> "1", "1.1–1.5" -> "1", "2-3" -> "2")
        # Prioritize explicitly detected current_chapter over inference from item_no
        if current_chapter:
            # Try to extract just the number from the chapter string if possible
            # e.g. "Chapter 2" -> "2", "2 Intro" -> "2"
            chapter_match = re.search(r'\d+', current_chapter)
            chapter_from_item = chapter_match.group(0) if chapter_match else current_chapter
        elif item_no:
            chapter_from_item = re.split(r'[.\-–]', item_no)[0]
        else:
            chapter_from_item = ""

        caption = para[match.end():].strip(" :.-")

        credit_line = ""

        # Same paragraph
        credit = extract_credit_sentence(para)
        if credit:
            credit_line = credit

        # Look ahead for credit lines after caption (up to 10 paragraphs)
        if not credit_line:
            for j in range(i + 1, min(i + 10, len(paragraphs))):
                next_p = paragraphs[j]
                if match_caption(next_p)[1] or CHAPTER_REGEX.match(next_p):
                    break
                credit = extract_credit_sentence(next_p)
                if credit:
                    credit_line = credit
                    # If we found a "Sources:" block, try to capture the following lines too
                    if "sources" in credit.lower():
                        source_lines = [credit]
                        # Capture subsequent reference lines (up to 5 more lines)
                        for k in range(j + 1, min(j + 5, len(paragraphs))):
                            next_ref = paragraphs[k]
                            # Stop at next caption or chapter
                            if match_caption(next_ref)[1] or CHAPTER_REGEX.match(next_ref):
                                break
                            # If it looks like a reference line (starts with capital or author name pattern)
                            if next_ref and (next_ref[0].isupper() or re.match(r'^[A-Z][a-z]+,', next_ref)):
                                source_lines.append(next_ref)
                            else:
                                break
                        if len(source_lines) > 1:
                            credit_line = " ".join(source_lines)
                    break

        # User requested to ignore empty captions and only include items with credit lines
        if caption and credit_line:
            # Clean caption: remove the credit line if it appears at the end
            # Normalize strings for comparison to handle whitespace differences
            norm_caption = " ".join(caption.split())
            norm_credit = " ".join(credit_line.split())
            
            if norm_credit in norm_caption:
                # Find index of credit in original caption to preserve original formatting of the rest
                # This is a simple approximation
                idx = caption.replace("\n", " ").find(credit_line.split()[0]) # Try to find start
                if idx != -1:
                     # Check if the rest roughly matches? 
                     # Simpler: just replace if exact match in normalized
                     pass

            # Robust replacement
            if credit_line in caption:
                caption = caption.replace(credit_line, "")
            else:
                # Fallback: check if caption ends with the credit text (ignoring potential paren differences)
                # e.g. caption ends in "Source.)" and credit is "Source."
                pass
            
            # Simple cleanup
            caption = caption.replace(credit_line, "") 
            caption = caption.strip(" .):")
            # If caption ends with open paren (leftover), remove it
            if caption.endswith("("):
                caption = caption[:-1].strip()

            results.append({
                "chapter": chapter_from_item,
                "item_type": item_type,
                "item_no": item_no,
                "caption": caption.strip(),
                "credit": credit_line,
                "needs_permission": needs_permission(caption, credit_line)
            })

    return results


# ======================================================
# EXCEL OUTPUT
# ======================================================

def write_permission_log(results, output_file):
    wb = Workbook()
    wb.remove(wb.active)

    wb.create_sheet("Sheet1")
    wb.create_sheet("Sheet2")
    log = wb.create_sheet("Permission Log")

    log["B1"] = "Enter Chap. (Fig/Table/Box)"
    log["C1"] = "#Choose Item Type"
    log["D1"] = "Enter Item #"
    log["G1"] = "Enter Figure Legend or Table/Box Title"
    log["H1"] = "Enter Credit Line from Chapter"
    log["I1"] = "Likely Needs Permission"

    log.column_dimensions['B'].width = 22
    log.column_dimensions['C'].width = 16
    log.column_dimensions['D'].width = 14
    log.column_dimensions['G'].width = 48
    log.column_dimensions['H'].width = 55
    log.column_dimensions['I'].width = 24

    row = 2
    for r in results:
        log[f"B{row}"] = r["chapter"]
        log[f"C{row}"] = r["item_type"]
        log[f"D{row}"] = r["item_no"]
        log[f"G{row}"] = r["caption"]
        log[f"H{row}"] = r["credit"]
        log[f"I{row}"] = r["needs_permission"]

        log[f"G{row}"].alignment = Alignment(wrap_text=True)
        log[f"H{row}"].alignment = Alignment(wrap_text=True)

        row += 1

    wb.save(output_file)
